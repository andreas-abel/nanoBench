#!/usr/bin/env python3

from collections import namedtuple
import xml.etree.ElementTree as ET
from xml.dom import minidom
import argparse
import re
from openpyxl import load_workbook

DocEntry = namedtuple('DocEntry', ['mnemonic', 'operands', 'ops', 'unit', 'lat', 'tp'])

def main():
   parser = argparse.ArgumentParser(description="Add data to XML file from AMD's doc")
   parser.add_argument('-xml')
   parser.add_argument('-xlsx')
   parser.add_argument('-outputXML')
   parser.add_argument('-arch')
   args = parser.parse_args()

   docEntrySet = set()
   mnemonicMap = dict()

   wb = load_workbook(args.xlsx, data_only=True)
   ws = wb.active
   for row in ws.iter_rows(min_row=2, values_only=True):
      mnemonicStr = row[0]
      if not mnemonicStr: continue

      if mnemonicStr.endswith('cc'):
         mnemonics = [mnemonicStr.replace('cc', p) for p in ['B', 'BE', 'L', 'LE', 'NB', 'NBE', 'NL', 'NLE', 'NO', 'NP', 'NS', 'NZ', 'O', 'P', 'S', 'Z']]
      else:
         mnemonics = mnemonicStr.replace(' (near)', '').replace('cc', '').split('/')

      for mnemonic in mnemonics:
         if mnemonic in ['AAA', 'AAD', 'AAM', 'AAS', 'ARPL', 'BOUND', 'DAA', 'DAS', 'INTO', 'JCXZ', 'LDS', 'LES','POPA', 'POPAD', 'POPD', 'POPFD', 'PUSHA', 'PUSHAD', 'PUSHFD']:
            # 32-bit instructions
            continue
         if mnemonic in ['CMPS', 'FCLEX', 'FINIT', 'FSAVE', 'FSTCW', 'FSTENV', 'FSTSW', 'INS', 'LODS', 'LOOPNZ', 'LOOPZ', 'MOVS', 'OUTS', 'PCLMULHQHQDQ', 'PCLMULHQLQDQ', 'PCLMULLQHQDQ', 'PCLMULLQLQDQ', 'RDPRU', 'SAL', 'SCAS', 'STOS', 'VGATHERDD', 'VGATHERDQ', 'VGATHERQD', 'VGATHERQQ','VPCLMULHQHQDQ', 'VPCLMULHQLQDQ', 'VPCLMULLQHQDQ', 'VPCLMULLQLQDQ', 'WAIT', 'XLATB']:
            # missing in XED
            continue
         if mnemonic in ['INT1', 'JECXZ']:
            # missing from XML file
            continue

         operands = row[1:5]

         ops = row[7]
         if ops == 'not supported':
            continue

         unit = row[8]
         lat = row[9]
         tp = row[10]

         if (ops is None) and (unit is None) and (lat is None) and (tp is None):
            continue

         de = DocEntry(mnemonic, operands, ops, unit, lat, tp)
         docEntrySet.add(de)
         mnemonicMap.setdefault(mnemonic, []).append(de)

   iclassAsmDict = dict()

   root = ET.parse(args.xml).getroot()
   for instrNode in root.iter('instruction'):
      if instrNode.attrib.get('evex', '') == '1':
         continue
      if instrNode.attrib['extension'] == 'VAES':
         continue
      iclass = instrNode.attrib['iclass']
      asm = instrNode.attrib['asm']
      iclassAsmDict.setdefault(iclass, set()).add(instrNode)
      iclassAsmDict.setdefault(re.sub('{.*} ', '', asm), set()).add(instrNode)

   #for x in set(op for de in docList for op in de.operands):
   #   print(x)

   xmlToDocDict = dict()

   for de in docEntrySet:
      if de.mnemonic not in iclassAsmDict:
         print('no XML entry found for ' + str(de))
         continue

      xmlFound = False
      for instrNode in iclassAsmDict[de.mnemonic]:
         explXmlOperands = [op for op in instrNode.findall('./operand') if not op.attrib.get('suppressed', '') == '1' and not op.attrib.get('implicit', '') == '1']
         docOperands = [op for op in de.operands if op is not None]

         if (not docOperands and any(op.attrib['type'] == 'mem' for op in explXmlOperands) and
             any(len(instrNode2.findall('./operand[@type="mem"]')) == 0 for instrNode2 in iclassAsmDict[de.mnemonic] if instrNode != instrNode2)):
            continue

         if docOperands and explXmlOperands and (len(explXmlOperands) != len(docOperands)):
            if any(len(explXmlOperands) == len([op for op in de2.operands if op is not None]) for de2 in mnemonicMap[de.mnemonic] if de!=de2):
               continue

         if docOperands and explXmlOperands:
            xmlOperands = explXmlOperands
         else:
            xmlOperands = [op for op in instrNode.findall('./operand')]

         invalid = False
         for docOp, xmlOp in zip(docOperands, xmlOperands):
            if de.mnemonic in ['CLZERO']: continue
            if xmlOp.attrib['type'] == 'mem' and set(de.operands) == {None}:
               invalid = True
               break
            if docOp is None: continue
            if docOp in ['pntr16/mem16:16/32']: continue

            if xmlOp.attrib['type'] == 'reg':
               if docOp == 'segmentReg':
                  if xmlOp.attrib.get('implicit', '') == '1': continue
               elif docOp in ['reg', 'reg/mem'] and xmlOp.attrib.get('implicit', '') != '1': continue
               elif not 'MM' in xmlOp.text:
                  if docOp == 'Sti' and xmlOp.text.startswith('ST'): continue
                  if docOp == 'ax' and xmlOp.text == 'AX': continue
                  if 'width' in xmlOp.attrib and re.search('reg(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue
               else:
                  if 'mmx' in docOp and xmlOp.text.startswith('MM'): continue
                  if 'xmm' in docOp and xmlOp.text.startswith('XMM'): continue
                  if 'ymm' in docOp and xmlOp.text.startswith('YMM'): continue
            elif xmlOp.attrib['type'] == 'mem':
               if docOp in ['mem', 'reg/mem', 'xmm2/mem', 'vm32x']: continue
               if re.search('mem(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue
            elif xmlOp.attrib['type'] in ['imm', 'relbr']:
               if docOp in ['imm', 'imm`', 'CL/Imm', 'xmm3/imm']: continue
               if re.search('imm(\d+/)*' + xmlOp.attrib['width'], docOp) is not None: continue

            invalid = True

         if invalid:
            continue

         if instrNode in xmlToDocDict:
            if (set(de.operands) != {None}) and (set(xmlToDocDict[instrNode].operands) == {None}):
               xmlFound = True
               xmlToDocDict[instrNode] = de
            elif (set(de.operands) == {None}) and (set(xmlToDocDict[instrNode].operands) != {None}):
               pass
            else:
               print('duplicate entry for ' + instrNode.attrib['string'] + ' found: ' + str(list(xmlToDocDict[instrNode])) + ', ' + str(list(de)))
         else:
            xmlFound = True
            xmlToDocDict[instrNode] = de

      if not xmlFound:
         print('no matching XML entry found for ' + str(de))

   print('Found data for ' + str(len(xmlToDocDict)) + ' instruction variants')

   for instrNode, de in xmlToDocDict.items():
      archNode = instrNode.find('./architecture[@name="{}"]'.format(args.arch))
      if archNode is None:
         archNode = ET.SubElement(instrNode, "architecture")
         archNode.attrib['name'] = args.arch

      docNode = ET.SubElement(archNode, "doc")
      if de.ops: docNode.attrib['uops'] = str(de.ops)
      if de.unit: docNode.attrib['ports'] = str(de.unit)
      if de.lat and de.lat != '-': docNode.attrib['latency'] = str(de.lat)
      if de.tp:
         try:
            if str(de.tp) == '0.33':
               docNode.attrib['TP'] = '3.00'
            else:
               docNode.attrib['TP'] = format(1/float(de.tp), '.2f')
         except ValueError:
            docNode.attrib['TP'] = de.tp

   with open(args.outputXML, "w") as f:
      rough_string = ET.tostring(root, 'utf-8')
      reparsed = minidom.parseString(rough_string)
      f.write('\n'.join([line for line in reparsed.toprettyxml(indent=' '*2).split('\n') if line.strip()]))


if __name__ == "__main__":
    main()
